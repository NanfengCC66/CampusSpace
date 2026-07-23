"""
数据导出服务
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone
from io import BytesIO


class ExportService:
    """数据导出服务"""
    
    @classmethod
    def export_bookings_to_excel(cls, bookings):
        """
        导出预约记录到Excel
        
        Args:
            bookings: 预约查询集
        
        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = '预约记录'
        
        # 设置表头
        headers = [
            '预约编号', '场地名称', '所属楼宇', '预约人', '使用目的',
            '预约用途', '开始时间', '结束时间', '参与人数', '状态',
            '联系人', '联系电话', '创建时间'
        ]
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据样式
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        # 填充数据
        status_map = {
            'pending': '待审批',
            'approved': '已通过',
            'rejected': '已拒绝',
            'cancelled': '已取消',
            'completed': '已完成',
            'expired': '已过期'
        }
        
        booking_type_map = {
            'teaching': '教学',
            'meeting': '会议',
            'activity': '活动',
            'self_study': '自习'
        }
        
        for row, booking in enumerate(bookings, 2):
            ws.cell(row=row, column=1, value=booking.booking_no).font = data_font
            ws.cell(row=row, column=2, value=booking.venue.name).font = data_font
            ws.cell(row=row, column=3, value=booking.venue.building.name).font = data_font
            ws.cell(row=row, column=4, value=booking.user.username).font = data_font
            ws.cell(row=row, column=5, value=booking.title).font = data_font
            ws.cell(row=row, column=6, value=booking_type_map.get(booking.booking_type, booking.booking_type)).font = data_font
            ws.cell(row=row, column=7, value=booking.start_time.strftime('%Y-%m-%d %H:%M')).font = data_font
            ws.cell(row=row, column=8, value=booking.end_time.strftime('%Y-%m-%d %H:%M')).font = data_font
            ws.cell(row=row, column=9, value=booking.participant_count).font = data_font
            ws.cell(row=row, column=10, value=status_map.get(booking.status, booking.status)).font = data_font
            ws.cell(row=row, column=11, value=booking.contact_name or '').font = data_font
            ws.cell(row=row, column=12, value=booking.contact_phone or '').font = data_font
            ws.cell(row=row, column=13, value=booking.created_at.strftime('%Y-%m-%d %H:%M')).font = data_font
            
            for col in range(1, 14):
                ws.cell(row=row, column=col).alignment = data_alignment
        
        # 设置列宽
        column_widths = [20, 20, 15, 15, 30, 10, 18, 18, 10, 10, 15, 15, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @classmethod
    def export_statistics_to_excel(cls, statistics_data, title='统计数据'):
        """
        导出统计数据到Excel
        
        Args:
            statistics_data: 统计数据字典
            title: 工作表标题
        
        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = title
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置数据样式
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        row = 1
        for key, value in statistics_data.items():
            # 写入标题
            cell = ws.cell(row=row, column=1, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # 写入值
            cell = ws.cell(row=row, column=2, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @classmethod
    def export_venue_usage_to_excel(cls, venue_usage_data):
        """
        导出场地使用统计到Excel
        
        Args:
            venue_usage_data: 场地使用数据列表
        
        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = '场地使用统计'
        
        # 设置表头
        headers = ['场地名称', '所属楼宇', '总预约数', '已完成', '待审批', '使用率']
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据样式
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        for row, data in enumerate(venue_usage_data, 2):
            ws.cell(row=row, column=1, value=data.get('venue_name', '')).font = data_font
            ws.cell(row=row, column=2, value=data.get('building_name', '')).font = data_font
            ws.cell(row=row, column=3, value=data.get('total_bookings', 0)).font = data_font
            ws.cell(row=row, column=4, value=data.get('completed', 0)).font = data_font
            ws.cell(row=row, column=5, value=data.get('pending', 0)).font = data_font
            ws.cell(row=row, column=6, value=f"{data.get('usage_rate', 0):.1%}").font = data_font
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = data_alignment
        
        # 设置列宽
        column_widths = [20, 15, 12, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output